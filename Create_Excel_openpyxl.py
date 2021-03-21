# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import pandas as pd
from openpyxl import Workbook
from datetime import datetime, timedelta

month = datetime.today().month
year = datetime.today().year
dt = str(datetime.today()-timedelta(days = 1)).split()[0]
filename = f'Efforts_{month}_{year}.xlsx'

# this creates an xlsx workbook with 4 worksheets
wb = Workbook()
ws = wb.active
ws.title = "Data"


ws['A1'] = 'Ticket No'
ws['B1'] = 'Ref No'
ws['C1'] = 'Corp ID'
ws['D1'] = 'Activity Type'
ws['E1'] = 'Activity'
ws['F1'] = 'Timecard Entry Date'
ws['G1'] = 'Effort'
ws['H1'] = 'Complexity'
ws['I1'] = 'AMorAD'
ws['J1'] = 'SOW'
ws['K1'] = 'Project'

tickets = ['','0','CHG0000000','CHG1111111','SCT1232121','INC1121212','SCT121221','CHG121212','INC1212122','INC1212122','INC121212']

l = len(tickets)-1
corp_id = 'rikushwa'
project = 'TOPSI'
complexity = 'Medium'
activity_type = ''
activity = ''
reference = ''
effort = 0
AMorAD = ''
SOW = ''
#date_format = workbook.add_format({'num_format' : 'yyyy-mm-dd hh:mm:ss'})

row = ws.max_row
column = ws.max_column

print(row,column)

ws.append(tickets)

my_dic = []

for ticket in tickets:
    if ticket == '':
        activity_type = 'Meetings / Communication'
        activity = 'Mail Communication'
        effort = 1.5
    elif ticket == '0':
        activity_type = 'Service-Task'
        activity = 'DSTUM'
        effort = 1.5
    elif ticket[:3] == 'CHG':
        activity_type = 'Change Request'
        activity = 'Third party coordination'
        effort = 1
    else:
        activity_type = 'Incident'
        activity = 'Incident'
        effort = 0.75

    ticket_details = ([ticket, reference, corp_id, activity_type, activity, dt, effort, complexity, AMorAD, SOW, project],)

    for tickets, reference, corp_id, activity_type, activity, date, effort, complexity, AMorAD, SOW, project in ticket_details:
        my_dic.append([ticket, reference, corp_id, activity_type, activity, dt, effort, complexity, AMorAD, SOW, project])

ws.append(my_dic)
wb.save(filename)

